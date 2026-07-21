"""Doldurulmuş KVKK anket çalışma kitabı (.xlsx) → süreçler + müvekkil profili.

VERBİS parse_inventory_xlsx'ten AYRI bir format: anket sayfaları
(NN-Departman) başlık adıyla eşlenir, 01-Genel sayfası etiket-eşleştirmeyle
profile çıkarır. Çıktı süreç şekli VERBİS parser ile aynıdır.
"""

from __future__ import annotations

import re
import unicodedata
import zipfile
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

_EXCLUDED_SHEETS = {"01-Genel-Sirket", "00-Dashboard", "15-Kanıt Belgeleri"}
_SKIP_STATUSES = {"Uygulanamaz", "Bilinmiyor"}

_LABEL_TO_FIELD = {
    "Ticari Unvan": "unvan",
    "MERSİS No": "mersis",
    "Vergi Dairesi": "vergi_dairesi",
    "Vergi No": "vergi_no",
    "Adres": "adres",
    "KEP Adresi": "kep",
    "E-posta": "eposta",
    "Telefon": "telefon",
}

class WorkbookImportError(Exception):
    pass


def _split(cell) -> list[str]:
    if cell is None:
        return []
    parts = re.split(r"[,;·\n]", str(cell))
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        n = unicodedata.normalize("NFC", p).strip()
        if not n or n in seen:
            continue
        seen.add(n)
        out.append(n)
    return out


def _is_department_sheet(name: str) -> bool:
    return bool(re.match(r"^\d{2}-", name)) and name not in _EXCLUDED_SHEETS


def _header_map(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if header_row is None:
        return {}
    return {
        str(value).strip(): idx
        for idx, value in enumerate(header_row)
        if value is not None and str(value).strip()
    }


def _looks_like_survey_workbook(wb) -> bool:
    has_genel = any(name.startswith("01-Genel") for name in wb.sheetnames)
    if not has_genel:
        return False
    for name in wb.sheetnames:
        if not _is_department_sheet(name):
            continue
        headers = _header_map(wb[name])
        if "İşlenen Kişisel Veri" in headers and "İlgili Kişi Grubu" in headers:
            return True
    return False


def _cell(row, headers: dict[str, int], label: str):
    idx = headers.get(label)
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _extract_processes(wb, sector: str) -> list[dict]:
    processes: list[dict] = []
    for name in wb.sheetnames:
        if not _is_department_sheet(name):
            continue
        sheet = wb[name]
        headers = _header_map(sheet)
        if "İşlenen Kişisel Veri" not in headers or "İlgili Kişi Grubu" not in headers:
            continue
        for row in sheet.iter_rows(min_row=3):
            durum = _cell(row, headers, "Durum")
            if durum is not None and str(durum).strip() in _SKIP_STATUSES:
                continue
            veri = _cell(row, headers, "İşlenen Kişisel Veri")
            kisi_grubu = _cell(row, headers, "İlgili Kişi Grubu")
            if not (str(veri or "").strip()) and not (str(kisi_grubu or "").strip()):
                continue

            yurtdisi = _cell(row, headers, "Yurtdışı Aktarım")
            aktarim = _split(_cell(row, headers, "Veri Alıcısı"))
            if str(yurtdisi or "").strip() == "Evet":
                aktarim = aktarim + ["Yurt dışına aktarım"]

            data = {
                "veri_turleri": _split(veri),
                "hukuki_sebepler": _split(_cell(row, headers, "Hukuki Sebep (KVKK → GDPR)")),
                "saklama_sureleri": _split(_cell(row, headers, "Saklama Süresi")),
                "toplama": _split(_cell(row, headers, "Veri Kaynağı")),
                "konum": _split(_cell(row, headers, "Kullanılan Sistem")),
                "aktarim": aktarim,
                "kategoriler": [],
                "amaclar": [],
                "dayanaklar": [],
                "islem": [],
                "ortam_format": [],
                "idari_tedbirler": [],
                "teknik_tedbirler": [],
            }

            bolum = _cell(row, headers, "Bölüm")
            soru = _cell(row, headers, "Soru / Süreç")
            processes.append({
                "sector": sector,
                "kisi_grubu": str(kisi_grubu or "").strip(),
                "departman": name,
                "is_sureci": str(bolum or "").strip() or "Genel",
                "alt_surec": str(soru or "").strip(),
                "data": data,
            })
    return processes


def _extract_profile(wb) -> dict:
    sheet_name = next((n for n in wb.sheetnames if n.startswith("01-Genel")), None)
    if sheet_name is None:
        return {}
    sheet = wb[sheet_name]
    profile: dict[str, str] = {}
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            label = str(cell.value).strip()
            field = _LABEL_TO_FIELD.get(label)
            if field is None or field in profile:
                continue
            if i + 1 >= len(row):
                continue
            value = row[i + 1].value
            value_str = str(value).strip() if value is not None else ""
            if value_str:
                profile[field] = value_str
    return profile


def parse_workbook_xlsx(content: bytes, sector: str) -> dict:
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError) as e:
        raise WorkbookImportError("Bozuk veya geçersiz .xlsx dosyası.") from e

    try:
        if not _looks_like_survey_workbook(wb):
            raise WorkbookImportError("Bu bir KVKK anket çalışma kitabı değil.")

        return {
            "processes": _extract_processes(wb, sector),
            "profile": _extract_profile(wb),
        }
    except WorkbookImportError:
        raise
    except Exception as e:
        raise WorkbookImportError("Çalışma kitabı beklenmeyen bir yapıda, ayrıştırılamadı.") from e
