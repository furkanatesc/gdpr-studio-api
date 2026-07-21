import io

import pytest
from openpyxl import Workbook

from app.workbook_import import WorkbookImportError, parse_workbook_xlsx

HEADER = [
    "No", "Bölüm", "Soru / Süreç", "Zorunlu?", "Cevap / Açıklama",
    "İşlenen Kişisel Veri", "İlgili Kişi Grubu", "Veri Kaynağı",
    "Kullanılan Sistem", "Veri Alıcısı", "Yurtdışı Aktarım",
    "Hukuki Sebep (KVKK → GDPR)", "Saklama Süresi", "Mevcut Doküman",
    "Kanıt Belgesi", "Risk / Not", "Sorumlu", "Durum",
]


def _build_workbook():
    wb = Workbook()
    genel = wb.active
    genel.title = "01-Genel-Sirket"
    genel["B5"] = "Ticari Unvan"
    genel["C5"] = "ACME A.Ş."
    genel["B7"] = "MERSİS No"
    genel["C7"] = "0123456789"

    ik = wb.create_sheet("02-İnsan Kaynakları")
    ik.append(["Sayfa Başlığı"])
    ik.append(HEADER)
    ik.append([
        1, "İşe Alım", "Özgeçmiş toplama", "Evet", "-",
        "Kimlik, İletişim", "Çalışan Adayı", "Kariyer Sitesi",
        "İK Sistemi", "", "Hayır",
        "Açık Rıza", "2 yıl", "-", "-", "-", "İK Md.", "Tamamlandı",
    ])
    ik.append([
        2, "Bordro", "Maaş ödemesi", "Evet", "-",
        "Kimlik, Banka Hesabı", "Çalışan", "Çalışandan",
        "Muhasebe Yazılımı", "Banka", "Evet",
        "Sözleşme", "10 yıl", "-", "-", "-", "Muhasebe Md.", "Tamamlandı",
    ])
    ik.append([
        3, "Genel", "Şirket kaç şubede faaliyet gösteriyor?", "Evet", "5",
        "", "", "", "", "", "",
        "", "", "-", "-", "-", "İK Md.", "Tamamlandı",
    ])
    ik.append([
        4, "İzin Takibi", "İzin talebi", "Hayır", "-",
        "Kimlik", "Çalışan", "Formdan", "-", "", "Hayır",
        "Sözleşme", "-", "-", "-", "-", "İK Md.", "Uygulanamaz",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_workbook_extracts_processes_and_profile():
    result = parse_workbook_xlsx(_build_workbook(), sector="sirket")

    processes = result["processes"]
    assert len(processes) == 2

    bordro = next(p for p in processes if p["alt_surec"] == "Maaş ödemesi")
    assert bordro["kisi_grubu"] == "Çalışan"
    assert bordro["departman"] == "02-İnsan Kaynakları"
    assert bordro["is_sureci"] == "Bordro"
    assert bordro["data"]["toplama"] == ["Çalışandan"]
    assert bordro["data"]["veri_turleri"] == ["Kimlik", "Banka Hesabı"]
    assert "Banka" in bordro["data"]["aktarim"]
    assert "Yurt dışına aktarım" in bordro["data"]["aktarim"]

    ise_alim = next(p for p in processes if p["alt_surec"] == "Özgeçmiş toplama")
    assert ise_alim["kisi_grubu"] == "Çalışan Adayı"
    assert "Yurt dışına aktarım" not in ise_alim["data"]["aktarim"]

    assert result["profile"] == {"unvan": "ACME A.Ş.", "mersis": "0123456789"}


def test_parse_workbook_rejects_non_survey_workbook():
    wb = Workbook()
    wb.active.title = "Sayfa1"
    wb.active["A1"] = "alakasız"
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(WorkbookImportError):
        parse_workbook_xlsx(buf.getvalue(), sector="sirket")


def test_parse_workbook_rejects_corrupt_bytes():
    with pytest.raises(WorkbookImportError):
        parse_workbook_xlsx(b"notxlsx", sector="sirket")


def test_parse_workbook_missing_header_row_raises_import_error():
    wb = Workbook()
    genel = wb.active
    genel.title = "01-Genel-Sirket"

    bozuk = wb.create_sheet("02-Bozuk Departman")
    bozuk.append(["Sayfa Başlığı"])

    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(WorkbookImportError):
        parse_workbook_xlsx(buf.getvalue(), sector="sirket")


def test_parse_workbook_skips_department_sheet_missing_header_row():
    wb = Workbook()
    genel = wb.active
    genel.title = "01-Genel-Sirket"

    bozuk = wb.create_sheet("02-Bozuk Departman")
    bozuk.append(["Sayfa Başlığı"])

    ik = wb.create_sheet("03-İnsan Kaynakları")
    ik.append(["Sayfa Başlığı"])
    ik.append(HEADER)
    ik.append([
        1, "İşe Alım", "Özgeçmiş toplama", "Evet", "-",
        "Kimlik, İletişim", "Çalışan Adayı", "Kariyer Sitesi",
        "İK Sistemi", "", "Hayır",
        "Açık Rıza", "2 yıl", "-", "-", "-", "İK Md.", "Tamamlandı",
    ])

    buf = io.BytesIO()
    wb.save(buf)

    result = parse_workbook_xlsx(buf.getvalue(), sector="sirket")
    processes = result["processes"]
    assert len(processes) == 1
    assert processes[0]["departman"] == "03-İnsan Kaynakları"
